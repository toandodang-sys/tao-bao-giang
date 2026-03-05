import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
import os
import re
from datetime import datetime, timedelta


# --- 1. LOGIC LÀM SẠCH FILE PPCT (Tránh lỗi dòng thừa) ---
def clean_ppct_dataframe(df):
    cols_str = " ".join([str(c).lower() for c in df.columns])
    if 'tiết' in cols_str and ('bài' in cols_str or 'nội dung' in cols_str):
        return df

    for i, row in df.head(10).iterrows():
        row_str = " ".join([str(x).lower() for x in row.values])
        if 'tiết' in row_str and ('bài' in row_str or 'nội dung' in row_str):
            df.columns = df.iloc[i]
            df = df.iloc[i + 1:].reset_index(drop=True)
            return df
    return df


# --- 2. LOGIC ĐỌC TKB TỪ GITHUB ---
def parse_school_tkb(df):
    teachers_tkb = {}
    current_teacher = None
    row_counter = 0

    for idx, row in df.iterrows():
        col0 = str(row.iloc[0]).strip()

        if col0 == 'nan' or col0 == '': continue
        if col0 == 'Tiết':
            row_counter = 0
            continue

        if col0.startswith('Tiết'):
            row_counter += 1
            buoi = "Sáng" if row_counter <= 5 else "Chiều"

            try:
                tiet_num = int(col0.replace('Tiết', '').strip())
            except:
                tiet_num = col0

            days = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6']
            for i, day in enumerate(days):
                if (i + 2) < len(row):
                    cell = str(row.iloc[i + 2]).strip()
                    if cell != 'nan' and cell != '' and '-' in cell:
                        parts = cell.split('-')
                        lop = parts[0].strip()
                        mon = parts[1].strip()

                        if current_teacher:
                            teachers_tkb[current_teacher].append({
                                'Thứ': day, 'Buổi': buoi, 'Tiết': tiet_num,
                                'Lớp': lop, 'Môn': mon
                            })
            continue

        current_teacher = col0
        if current_teacher not in teachers_tkb:
            teachers_tkb[current_teacher] = []

    return teachers_tkb


def load_saved_tkb():
    """Tự động tìm và tải file TKB trực tiếp từ thư mục mã nguồn (GitHub)"""
    for ext in ['xlsx', 'xls', 'csv']:
        filename = f"tkb_truong.{ext}"
        if os.path.exists(filename):
            try:
                if ext == 'csv':
                    df = pd.read_csv(filename, header=None)
                else:
                    df = pd.read_excel(filename, header=None)
                return parse_school_tkb(df)
            except Exception as e:
                st.error(f"Lỗi đọc file {filename}: {e}")
                return None
    return None


# --- 3. HÀM TÌM TÊN BÀI TỪ PPCT ---
def find_lesson_name(df_ppct, khoi, tiet_ppct, mon_tkb_clean):
    """Hàm hỗ trợ tìm tên bài dựa vào Khối, Tiết và Môn đã được chuẩn hóa"""
    try:
        query = (df_ppct['Tiết PPCT'] == tiet_ppct)

        # Bộ lọc Khối (Nếu PPCT áp dụng cho mọi khối thì Khối_Clean là rỗng)
        mask_khoi = (df_ppct['Khối_Clean'] == khoi) | (df_ppct['Khối_Clean'] == "") | df_ppct['Khối_Clean'].isna()
        query = query & mask_khoi

        # Bộ lọc Môn
        mask_mon = df_ppct['Môn_Clean'].apply(
            lambda x: mon_tkb_clean in str(x) or str(x) in mon_tkb_clean or
                      (mon_tkb_clean == 'tin' and str(x) == 'tin học') or
                      (mon_tkb_clean == 'tin học' and str(x) == 'tin')
        )
        query = query & mask_mon

        match = df_ppct[query]

        if not match.empty:
            return str(match.iloc[0]['Tên bài']).strip()
        else:
            return f"⚠️ Không tìm thấy bài (Khối {khoi} - Tiết {tiet_ppct})"
    except Exception as e:
        return "⚠️ Lỗi xử lý PPCT"


# --- HÀM ĐỌC SỐ THÀNH CHỮ (0-99) ---
def number_to_words_vn(n):
    if n == 0: return "không"
    units = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
    if n < 10: return units[n]
    if n == 10: return "mười"
    if n < 20:
        if n == 15: return "mười lăm"
        return "mười " + units[n % 10]

    tens = units[n // 10] + " mươi"
    ones_digit = n % 10

    if ones_digit == 0: return tens
    if ones_digit == 1: return tens + " mốt"
    if ones_digit == 4: return tens + " tư"
    if ones_digit == 5: return tens + " lăm"
    return tens + " " + units[ones_digit]


# --- 4. LOGIC TẠO FILE EXCEL BÁO CÁO ---
def create_excel_report(teacher_name, chuc_vu, to_chuyen_mon, nam_hoc, hoc_ky, week_num, start_date, end_date,
                        loai_kiem_nhiem, kiem_nhiem, loai_kiem_nhiem_2, kiem_nhiem_2, report_data):
    output_path = f"Bao_Cao_Tuan_{week_num}_{teacher_name.replace(' ', '_')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo Cáo"

    # ---------------- THIẾT LẬP TRANG IN ----------------
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Giấy ngang
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # Khổ A4
    ws.page_setup.fitToWidth = 1  # Vừa với 1 trang bề ngang
    ws.page_setup.fitToHeight = 0  # Chiều dọc tự do nhảy trang

    # Căn giữa trang in theo chiều ngang
    ws.print_options.horizontalCentered = True

    # Chỉnh lề (Margins) nhỏ lại để chứa được nhiều nội dung
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.3, footer=0.3)

    # Định dạng chung
    font_normal = Font(name='Times New Roman', size=12)
    font_bold = Font(name='Times New Roman', size=12, bold=True)
    font_italic = Font(name='Times New Roman', size=12, italic=True)
    font_title = Font(name='Times New Roman', size=14, bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ---------------- HEADER ----------------
    ws.merge_cells('C1:F1')
    ws['C1'] = "UBND XÃ BA TƠ"
    ws.merge_cells('C2:F2')
    ws['C2'] = "TRƯỜNG THCS BA TƠ"
    ws['C1'].font = font_bold
    ws['C2'].font = font_bold
    ws['C1'].alignment = align_center
    ws['C2'].alignment = align_center

    ws.merge_cells('I1:N1')
    ws['I1'] = "CỘNG HOÀ XÃ HỘI CHỦ NGHĨA VIỆT NAM"
    ws.merge_cells('I2:N2')
    ws['I2'] = "Độc lập – Tự do – Hạnh phúc"
    ws['I1'].font = font_bold
    ws['I2'].font = font_bold
    ws['I1'].alignment = align_center
    ws['I2'].alignment = align_center

    # ---------------- TITLE ----------------
    ws.merge_cells('A4:N4')
    ws[
        'A4'] = f"BÁO CÁO THỰC HIỆN KẾ HOẠCH DẠY HỌC, TIẾN ĐỘ CHƯƠNG TRÌNH, SỐ TIẾT DẠY TUẦN {week_num}, NĂM HỌC {nam_hoc.upper()}"
    ws['A4'].font = font_title
    ws['A4'].alignment = align_center

    ws.merge_cells('A5:N5')
    start_str = start_date.strftime("%d/%m/%Y")
    end_str = end_date.strftime("%d/%m/%Y")
    ws['A5'] = f"(Tuần: {week_num}. Từ ngày {start_str} đến ngày {end_str}, học kỳ {hoc_ky}, năm học {nam_hoc})"
    ws['A5'].font = font_italic
    ws['A5'].alignment = align_center

    # ---------------- THÔNG TIN GIÁO VIÊN ĐỘNG ----------------
    ws.merge_cells('C6:N6')
    ws['C6'] = f"Họ và tên: {teacher_name}, Chức vụ: {chuc_vu}; Tổ CM: {to_chuyen_mon}."
    ws['C6'].font = font_bold

    ws.merge_cells('C7:N7')

    # Logic tính toán Tự động lấy Môn và Lớp dạy (loại trừ Chào cờ, SHL)
    subject_class_map = {}
    for item in report_data:
        mon_tkb = str(item.get('Môn', '')).strip()
        lop_tkb = str(item.get('Lớp', '')).strip()
        mon_lower = mon_tkb.lower()

        # Loại trừ môn Chào cờ và SHL/Sinh hoạt lớp
        if 'chào cờ' in mon_lower or 'shl' in mon_lower or 'sinh hoạt' in mon_lower:
            continue

        if mon_tkb and lop_tkb:
            if mon_tkb not in subject_class_map:
                subject_class_map[mon_tkb] = set()
            subject_class_map[mon_tkb].add(lop_tkb)

    day_mon_str_parts = []
    for m, l_set in subject_class_map.items():
        # Sắp xếp tên lớp cho đẹp (VD: 6A1, 6A2)
        sorted_lops = sorted(list(l_set))
        day_mon_str_parts.append(f"{m} lớp {', '.join(sorted_lops)}")

    if day_mon_str_parts:
        ws['C7'] = f"Dạy môn, lớp: {'; '.join(day_mon_str_parts)}"
    else:
        ws[
            'C7'] = "Dạy môn, lớp: ......................................................................................................."

    ws['C7'].font = font_normal

    ws.merge_cells('C8:N8')
    ws['C8'] = "Chủ nhiệm lớp:….........   Số tiết CN/tuần: ...... tiết/tuần"
    ws['C8'].font = font_normal

    # Kiêm nhiệm 1
    ws.merge_cells('C9:N9')
    txt_loai_kn = loai_kiem_nhiem if loai_kiem_nhiem.strip() != "" else "..................."
    txt_so_tiet_kn = kiem_nhiem if kiem_nhiem > 0 else "......"

    has_kn2 = (loai_kiem_nhiem_2.strip() != "" or kiem_nhiem_2 > 0)
    label_kn1 = "Kiêm nhiệm/chức vụ 1" if has_kn2 else "Kiêm nhiệm/chức vụ"

    ws['C9'] = f"{label_kn1}: {txt_loai_kn}   Số tiết/tuần: {txt_so_tiet_kn}/tuần"
    ws['C9'].font = font_normal

    current_row = 10

    # Kiêm nhiệm 2 (Tự động sinh dòng nếu có)
    if has_kn2:
        ws.merge_cells(f'C{current_row}:N{current_row}')
        txt_loai_kn_2 = loai_kiem_nhiem_2 if loai_kiem_nhiem_2.strip() != "" else "..................."
        txt_so_tiet_kn_2 = kiem_nhiem_2 if kiem_nhiem_2 > 0 else "......"
        ws[f'C{current_row}'] = f"Kiêm nhiệm/chức vụ 2: {txt_loai_kn_2}   Số tiết/tuần: {txt_so_tiet_kn_2}/tuần"
        ws[f'C{current_row}'].font = font_normal
        current_row += 1

    # Tổng số tiết
    ws.merge_cells(f'C{current_row}:N{current_row}')
    total_kiem_nhiem = kiem_nhiem + kiem_nhiem_2
    ws[
        f'C{current_row}'] = f"Tổng số tiết giảng dạy: {len(report_data)} tiết; tổng số tiết kiêm nhiệm: {total_kiem_nhiem}. Cộng tổng số tiết/tuần: {len(report_data) + total_kiem_nhiem}."
    ws[f'C{current_row}'].font = font_normal
    current_row += 1

    # Định mức
    ws[f'C{current_row}'] = "Số tiết định mức:"
    ws[f'C{current_row}'].font = font_normal
    ws[f'E{current_row}'] = 19
    ws[f'E{current_row}'].font = font_normal
    ws[f'E{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'F{current_row}'] = "tiết/tuần"
    ws[f'F{current_row}'].font = font_normal
    current_row += 1

    # ---------------- BẢNG DỮ LIỆU CHÍNH ----------------
    start_row = current_row + 1
    headers = [
        "TT", "Thứ,\n ngày", "Buổi", "Tiết", "Dạy môn,\nkiêm nhiệm", "Lớp",
        "Tiết theo\nPPCT", "Tổng số tiết\nthực dạy,\nkiêm nhiệm", "Số tiết\nđi công tác",
        "Số tiết\ndạy thay", "Số tiết lấp giờ,\ntăng tiết, bù", "Số tiết coi KT,\ndự giờ thi GVDG,\nBD, PĐ",
        "Tên bài / Chủ đề / Nội dung", "Ghi chú"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = thin_border

    # Căn chỉnh độ rộng cột
    widths = {
        'A': 5, 'B': 13, 'C': 8, 'D': 6, 'E': 14, 'F': 8,
        'G': 10, 'H': 13, 'I': 10, 'J': 10, 'K': 13, 'L': 15, 'M': 45, 'N': 15
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Điền dữ liệu và đếm số loại tiết
    counts = {8: 0, 9: 0, 10: 0, 11: 0, 12: 0}

    # Bản đồ tính ngày dựa trên start_date đã chọn
    monday_date = start_date - timedelta(days=start_date.weekday())
    day_offset_map = {'Thứ 2': 0, 'Thứ 3': 1, 'Thứ 4': 2, 'Thứ 5': 3, 'Thứ 6': 4, 'Thứ 7': 5, 'Chủ Nhật': 6}

    for i, row_data in enumerate(report_data):
        curr = start_row + 1 + i

        for col_num in range(1, 15):
            cell = ws.cell(row=curr, column=col_num)
            cell.border = thin_border
            cell.font = font_normal
            if col_num in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
                cell.alignment = align_center
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        day_str = row_data['Thứ']
        actual_date = monday_date + timedelta(days=day_offset_map.get(day_str, 0))
        thu_ngay_str = f"{day_str}\n{actual_date.strftime('%d/%m/%Y')}"

        ws.cell(row=curr, column=1).value = i + 1
        ws.cell(row=curr, column=2).value = thu_ngay_str
        ws.cell(row=curr, column=3).value = row_data['Buổi']
        ws.cell(row=curr, column=4).value = row_data['Tiết']
        ws.cell(row=curr, column=5).value = row_data['Môn']
        ws.cell(row=curr, column=6).value = row_data['Lớp']
        ws.cell(row=curr, column=7).value = row_data['Tiết PPCT']
        ws.cell(row=curr, column=13).value = row_data['Tên Bài']

        loai_tiet = row_data.get('Loại Tiết', 'Thực dạy / Kiêm nhiệm')

        if loai_tiet != 'Thực dạy / Kiêm nhiệm':
            ws.cell(row=curr, column=14).value = loai_tiet

        if loai_tiet == 'Thực dạy / Kiêm nhiệm':
            ws.cell(row=curr, column=8).value = 1
            counts[8] += 1
        elif loai_tiet == 'Đi công tác':
            ws.cell(row=curr, column=9).value = 1
            counts[9] += 1
        elif loai_tiet == 'Dạy thay':
            ws.cell(row=curr, column=10).value = 1
            counts[10] += 1
        elif loai_tiet == 'Lấp giờ, tăng tiết, bù':
            ws.cell(row=curr, column=11).value = 1
            counts[11] += 1
        elif loai_tiet == 'Coi KT, dự giờ, BD, PĐ':
            ws.cell(row=curr, column=12).value = 1
            counts[12] += 1

    # ---------------- XỬ LÝ GỘP Ô (MERGE CELLS) CHO THỨ VÀ BUỔI ----------------
    data_start_row = start_row + 1
    if len(report_data) > 1:
        # Biến theo dõi gộp cột Thứ (Cột B / column 2)
        start_merge_thu = data_start_row
        current_thu = report_data[0]['Thứ']

        # Biến theo dõi gộp cột Buổi (Cột C / column 3)
        start_merge_buoi = data_start_row
        current_buoi = (report_data[0]['Thứ'], report_data[0]['Buổi'])

        for i in range(1, len(report_data)):
            row_idx = data_start_row + i

            # Xử lý gộp Thứ
            if report_data[i]['Thứ'] != current_thu:
                if start_merge_thu < row_idx - 1:
                    ws.merge_cells(start_row=start_merge_thu, start_column=2, end_row=row_idx - 1, end_column=2)
                current_thu = report_data[i]['Thứ']
                start_merge_thu = row_idx

            # Xử lý gộp Buổi (phải trùng cả Thứ và Buổi thì mới gộp)
            buoi_key = (report_data[i]['Thứ'], report_data[i]['Buổi'])
            if buoi_key != current_buoi:
                if start_merge_buoi < row_idx - 1:
                    ws.merge_cells(start_row=start_merge_buoi, start_column=3, end_row=row_idx - 1, end_column=3)
                current_buoi = buoi_key
                start_merge_buoi = row_idx

        # Lệnh gộp cho nhóm cuối cùng trong danh sách
        last_row_idx = data_start_row + len(report_data) - 1
        if start_merge_thu < last_row_idx:
            ws.merge_cells(start_row=start_merge_thu, start_column=2, end_row=last_row_idx, end_column=2)
        if start_merge_buoi < last_row_idx:
            ws.merge_cells(start_row=start_merge_buoi, start_column=3, end_row=last_row_idx, end_column=3)

    # ---------------- FOOTER & THỐNG KÊ ĐỘNG ----------------
    sum_row = start_row + len(report_data) + 1
    total_tiet = sum(counts.values()) + total_kiem_nhiem
    thua = total_tiet - 19 if total_tiet > 19 else 0
    thieu = 19 - total_tiet if total_tiet < 19 else 0

    # Tạo chuỗi "Bằng chữ" cho thừa và thiếu
    str_thua = f"Thừa {number_to_words_vn(thua)} tiết" if thua > 0 else "................................"
    str_thieu = f"Thiếu {number_to_words_vn(thieu)} tiết" if thieu > 0 else "................................"

    def write_summary_row(row, text, val_col, val, is_bold=False):
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = text
        ws[f'B{row}'].font = font_bold if is_bold else font_normal
        ws[f'H{row}'] = val
        ws[f'H{row}'].font = font_bold if is_bold else font_normal
        ws[f'I{row}'] = "tiết."
        ws[f'I{row}'].font = font_normal

    write_summary_row(sum_row, "Cộng tổng số tiết dạy:", "H", len(report_data), is_bold=True)
    sum_row += 1

    # Kiêm nhiệm 1
    ten_kiem_nhiem_hien_thi_1 = f"Số tiết kiêm nhiệm {loai_kiem_nhiem.strip()}:" if loai_kiem_nhiem.strip() else "Số tiết kiêm nhiệm:"
    write_summary_row(sum_row, ten_kiem_nhiem_hien_thi_1, "H", kiem_nhiem, is_bold=True)
    sum_row += 1

    # Sinh thêm dòng Kiêm nhiệm 2 nếu có
    if has_kn2:
        ten_kiem_nhiem_hien_thi_2 = f"Số tiết kiêm nhiệm {loai_kiem_nhiem_2.strip()}:" if loai_kiem_nhiem_2.strip() else "Số tiết kiêm nhiệm 2:"
        write_summary_row(sum_row, ten_kiem_nhiem_hien_thi_2, "H", kiem_nhiem_2, is_bold=True)
        sum_row += 1

    write_summary_row(sum_row, "Số tiết thực dạy, kiêm nhiệm (1):", "H", counts[8] + total_kiem_nhiem)
    sum_row += 1
    write_summary_row(sum_row, "Số tiết đi công tác (2):", "H", counts[9])
    sum_row += 1
    write_summary_row(sum_row, "Số tiết dạy thay (3):", "H", counts[10])
    sum_row += 1
    write_summary_row(sum_row, "Số tiết lấp giờ, tăng tiết, bù (4):", "H", counts[11])
    sum_row += 1
    write_summary_row(sum_row, "Số tiết coi KT, dự giờ thi GVDG, BD, PĐ (5):", "H", counts[12])
    sum_row += 1

    write_summary_row(sum_row, "Tổng cộng số tiết (1+2+3+4+5):", "H", total_tiet, is_bold=True)
    sum_row += 1

    ws.merge_cells(f'D{sum_row}:G{sum_row}')
    ws[f'D{sum_row}'] = "- Số tiết thừa:"
    ws[f'D{sum_row}'].font = font_normal
    ws[f'H{sum_row}'] = thua
    ws[f'H{sum_row}'].font = font_normal
    ws[f'I{sum_row}'] = "tiết."
    ws[f'I{sum_row}'].font = font_normal
    ws.merge_cells(f'L{sum_row}:N{sum_row}')
    ws[f'L{sum_row}'] = f"Bằng chữ: {str_thua}"
    ws[f'L{sum_row}'].font = font_normal
    sum_row += 1

    ws.merge_cells(f'D{sum_row}:G{sum_row}')
    ws[f'D{sum_row}'] = "- Số tiết thiếu:"
    ws[f'D{sum_row}'].font = font_normal
    ws[f'H{sum_row}'] = thieu
    ws[f'H{sum_row}'].font = font_normal
    ws[f'I{sum_row}'] = "tiết."
    ws[f'I{sum_row}'].font = font_normal
    ws.merge_cells(f'L{sum_row}:N{sum_row}')
    ws[f'L{sum_row}'] = f"Bằng chữ: {str_thieu}"
    ws[f'L{sum_row}'].font = font_normal
    sum_row += 1

    # Chữ ký Người lập - Ngày tháng năm tự động theo thời gian thực
    ws.merge_cells(f'L{sum_row}:N{sum_row}')
    today = datetime.now()
    ws[f'L{sum_row}'] = f"Ba Tơ, ngày {today.day:02d} tháng {today.month:02d} năm {today.year}"
    ws[f'L{sum_row}'].alignment = Alignment(horizontal='center')
    ws[f'L{sum_row}'].font = font_italic
    sum_row += 1

    ws.merge_cells(f'L{sum_row}:N{sum_row}')
    ws[f'L{sum_row}'] = "Người lập"
    ws[f'L{sum_row}'].alignment = Alignment(horizontal='center')
    ws[f'L{sum_row}'].font = font_bold
    sum_row += 4

    ws.merge_cells(f'L{sum_row}:N{sum_row}')
    ws[f'L{sum_row}'] = teacher_name
    ws[f'L{sum_row}'].alignment = Alignment(horizontal='center')
    ws[f'L{sum_row}'].font = font_bold

    # Kích hoạt ngắt trang (Page Break) mềm mại tại đây nếu bảng quá dài
    if sum_row > 45:
        ws.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=sum_row))

    # ---------------- PHẦN DUYỆT CỦA TỔ CHUYÊN MÔN ----------------
    row_duyet = sum_row + 2
    ws.merge_cells(f'D{row_duyet}:H{row_duyet}')
    ws[f'D{row_duyet}'] = "PHẦN DUYỆT CỦA TỔ CHUYÊN MÔN"
    ws[f'D{row_duyet}'].font = font_bold
    ws.merge_cells(f'J{row_duyet}:M{row_duyet}')
    ws[f'J{row_duyet}'] = f"(GV: {teacher_name})"
    ws[f'J{row_duyet}'].font = font_normal

    def write_to_row(r, text, val):
        ws.merge_cells(f'D{r}:G{r}')
        ws[f'D{r}'] = text
        ws[f'D{r}'].font = font_normal
        ws[f'I{r}'] = val
        ws[f'I{r}'].font = font_normal
        ws[f'I{r}'].alignment = Alignment(horizontal='right')
        ws[f'J{r}'] = "tiết"
        ws[f'J{r}'].font = font_normal

    write_to_row(row_duyet + 1, "- Số tiết dạy, kiêm nhiệm:", counts[8] + total_kiem_nhiem)
    write_to_row(row_duyet + 2, "- Số tiết đi công tác:", counts[9])
    write_to_row(row_duyet + 3, "- Số tiết dạy thay:", counts[10])
    write_to_row(row_duyet + 4, "- Số tiết dạy lấp giờ, tăng tiết, bù:", counts[11])
    write_to_row(row_duyet + 5, "- Số tiết coi KT,DG GVDG, BD, PĐ:", counts[12])

    ws.merge_cells(f'E{row_duyet + 6}:G{row_duyet + 6}')
    ws[f'E{row_duyet + 6}'] = "Tổng cộng:"
    ws[f'E{row_duyet + 6}'].font = font_bold
    ws[f'I{row_duyet + 6}'] = total_tiet
    ws[f'I{row_duyet + 6}'].font = font_bold
    ws[f'I{row_duyet + 6}'].alignment = Alignment(horizontal='right')
    ws[f'J{row_duyet + 6}'] = "tiết"
    ws[f'J{row_duyet + 6}'].font = font_normal

    ws.merge_cells(f'E{row_duyet + 7}:G{row_duyet + 7}')
    ws[f'E{row_duyet + 7}'] = "- Số tiết thừa:"
    ws[f'E{row_duyet + 7}'].font = font_normal
    ws[f'I{row_duyet + 7}'] = thua
    ws[f'I{row_duyet + 7}'].font = font_normal
    ws[f'I{row_duyet + 7}'].alignment = Alignment(horizontal='right')
    ws[f'J{row_duyet + 7}'] = "tiết"
    ws[f'J{row_duyet + 7}'].font = font_normal
    ws.merge_cells(f'K{row_duyet + 7}:M{row_duyet + 7}')
    ws[f'K{row_duyet + 7}'] = f"Bằng chữ: {str_thua}"
    ws[f'K{row_duyet + 7}'].font = font_normal

    ws.merge_cells(f'E{row_duyet + 8}:G{row_duyet + 8}')
    ws[f'E{row_duyet + 8}'] = "- Số tiết thiếu:"
    ws[f'E{row_duyet + 8}'].font = font_normal
    ws[f'I{row_duyet + 8}'] = thieu
    ws[f'I{row_duyet + 8}'].font = font_normal
    ws[f'I{row_duyet + 8}'].alignment = Alignment(horizontal='right')
    ws[f'J{row_duyet + 8}'] = "tiết"
    ws[f'J{row_duyet + 8}'].font = font_normal
    ws.merge_cells(f'K{row_duyet + 8}:M{row_duyet + 8}')
    ws[f'K{row_duyet + 8}'] = f"Bằng chữ: {str_thieu}"
    ws[f'K{row_duyet + 8}'].font = font_normal

    ws.merge_cells(f'E{row_duyet + 9}:G{row_duyet + 9}')
    ws[f'E{row_duyet + 9}'] = "TỔ TRƯỞNG"
    ws[f'E{row_duyet + 9}'].font = font_bold
    ws[f'E{row_duyet + 9}'].alignment = Alignment(horizontal='center')

    # ---------------- PHẦN DUYỆT CỦA CHUYÊN MÔN TRƯỜNG ----------------
    row_cm_truong = row_duyet + 14
    ws.merge_cells(f'D{row_cm_truong}:H{row_cm_truong}')
    ws[f'D{row_cm_truong}'] = "PHẦN DUYỆT CỦA CHUYÊN MÔN TRƯỜNG"
    ws[f'D{row_cm_truong}'].font = font_bold

    write_to_row(row_cm_truong + 1, "- Số tiết dạy, kiêm nhiệm:", counts[8] + total_kiem_nhiem)
    write_to_row(row_cm_truong + 2, "- Số tiết đi công tác:", counts[9])
    write_to_row(row_cm_truong + 3, "- Số tiết dạy thay:", counts[10])
    write_to_row(row_cm_truong + 4, "- Số tiết dạy lấp giờ, tăng tiết, bù:", counts[11])
    write_to_row(row_cm_truong + 5, "- Số tiết coi KT,DG GVDG, BD, PĐ:", counts[12])

    ws.merge_cells(f'E{row_cm_truong + 6}:G{row_cm_truong + 6}')
    ws[f'E{row_cm_truong + 6}'] = "Tổng cộng:"
    ws[f'E{row_cm_truong + 6}'].font = font_bold
    ws[f'I{row_cm_truong + 6}'] = total_tiet
    ws[f'I{row_cm_truong + 6}'].font = font_bold
    ws[f'I{row_cm_truong + 6}'].alignment = Alignment(horizontal='right')
    ws[f'J{row_cm_truong + 6}'] = "tiết"
    ws[f'J{row_cm_truong + 6}'].font = font_normal

    ws.merge_cells(f'E{row_cm_truong + 7}:G{row_cm_truong + 7}')
    ws[f'E{row_cm_truong + 7}'] = "- Số tiết thừa:"
    ws[f'E{row_cm_truong + 7}'].font = font_normal
    ws[f'I{row_cm_truong + 7}'] = thua
    ws[f'I{row_cm_truong + 7}'].font = font_normal
    ws[f'I{row_cm_truong + 7}'].alignment = Alignment(horizontal='right')
    ws[f'J{row_cm_truong + 7}'] = "tiết"
    ws[f'J{row_cm_truong + 7}'].font = font_normal
    ws.merge_cells(f'K{row_cm_truong + 7}:M{row_cm_truong + 7}')
    ws[f'K{row_cm_truong + 7}'] = f"Bằng chữ: {str_thua}"
    ws[f'K{row_cm_truong + 7}'].font = font_normal

    ws.merge_cells(f'E{row_cm_truong + 8}:G{row_cm_truong + 8}')
    ws[f'E{row_cm_truong + 8}'] = "- Số tiết thiếu:"
    ws[f'E{row_cm_truong + 8}'].font = font_normal
    ws[f'I{row_cm_truong + 8}'] = thieu
    ws[f'I{row_cm_truong + 8}'].font = font_normal
    ws[f'I{row_cm_truong + 8}'].alignment = Alignment(horizontal='right')
    ws[f'J{row_cm_truong + 8}'] = "tiết"
    ws[f'J{row_cm_truong + 8}'].font = font_normal
    ws.merge_cells(f'K{row_cm_truong + 8}:M{row_cm_truong + 8}')
    ws[f'K{row_cm_truong + 8}'] = f"Bằng chữ: {str_thieu}"
    ws[f'K{row_cm_truong + 8}'].font = font_normal

    ws.merge_cells(f'E{row_cm_truong + 9}:G{row_cm_truong + 9}')
    ws[f'E{row_cm_truong + 9}'] = "CHUYÊN MÔN TRƯỜNG"
    ws[f'E{row_cm_truong + 9}'].font = font_bold
    ws[f'E{row_cm_truong + 9}'].alignment = Alignment(horizontal='center')

    wb.save(output_path)
    return output_path


# --- 5. GIAO DIỆN STREAMLIT WEB APP ---
st.set_page_config(page_title="Hệ thống Báo Cáo Giáo Viên", layout="wide", page_icon="☀️")
st.title("☀️ Cổng Tự Động Hóa Báo Cáo Giảng Dạy Trường THCS Ba Tơ")

# Tự động đọc TKB từ thư mục GitHub (Không cần Admin up)
teachers_dict = load_saved_tkb()

if not teachers_dict:
    st.info(
        "👋 Hệ thống chưa tìm thấy dữ liệu Thời khóa biểu. Quản trị viên vui lòng tải file TKB lên GitHub và đặt tên là 'tkb_truong.xlsx' hoặc 'tkb_truong.csv' để bắt đầu sử dụng.")
else:
    list_gv = list(teachers_dict.keys())

    # Trích xuất danh sách Lớp và Môn tự động từ TKB toàn trường
    all_lops = set()
    all_mons = set()
    for gv, lessons in teachers_dict.items():
        for ls in lessons:
            all_lops.add(str(ls['Lớp']))
            all_mons.add(str(ls['Môn']))
    list_all_lops = sorted(list(all_lops)) if all_lops else ["6A1", "7A1", "8A1", "9A1"]
    list_all_mons = sorted(list(all_mons)) if all_mons else ["Tin học", "Toán", "Ngữ văn"]

    st.success(f"✅ Đã tải thành công TKB của **{len(list_gv)}** giáo viên trực tiếp từ hệ thống.")

    st.markdown("### 👨‍🏫 Thông tin Giáo viên")

    row1_col1, row1_col2, row1_col3 = st.columns([1, 1, 1])
    with row1_col1:
        selected_teacher = st.selectbox("Chọn tên của bạn:", ["-- Chọn giáo viên --"] + list_gv)
    with row1_col2:
        chuc_vu = st.selectbox("Chức vụ:", ["Giáo viên", "Tổ trưởng chuyên môn", "Tổ phó chuyên môn", "Hiệu trưởng",
                                            "Phó hiệu trưởng"])
    with row1_col3:
        to_chuyen_mon = st.selectbox("Tổ Chuyên môn:", ["Khoa học Tự nhiên", "Khoa học Xã hội"])

    st.markdown("### 📊 Nguồn dữ liệu của bạn")
    up_ppct = st.file_uploader("Tải lên các file PPCT cá nhân (Có thể quét chọn nhiều file cùng lúc)",
                               type=["xlsx", "csv"], accept_multiple_files=True)

    st.markdown("**Cấu hình thời gian & Năm học:**")
    time_col1, time_col2, time_col3, time_col4, time_col5 = st.columns(5)
    today_date = datetime.today().date()
    default_start = today_date - timedelta(days=today_date.weekday())  # Mặc định đầu tuần Thứ 2

    with time_col1:
        nam_hoc = st.text_input("Năm học:", value="2025 - 2026")
    with time_col2:
        hoc_ky = st.selectbox("Học kỳ:", ["I", "II", "Hè"], index=1)
    with time_col3:
        selected_week = st.number_input("📅 Số Tuần:", min_value=1, value=1, step=1)
    with time_col4:
        start_date = st.date_input("🗓️ Từ ngày:", value=default_start, format="DD/MM/YYYY")
    with time_col5:
        end_date = st.date_input("🗓️ Đến ngày:", value=default_start + timedelta(days=5), format="DD/MM/YYYY")

    st.markdown("**Bổ sung thông tin kiêm nhiệm (Nếu có):**")
    kn_col1, kn_col2, kn_col3, kn_col4 = st.columns(4)
    with kn_col1:
        loai_kiem_nhiem = st.text_input("📝 Kiêm nhiệm 1:", placeholder="VD: CNTT, Tổ trưởng...")
    with kn_col2:
        kiem_nhiem = st.number_input("⏱️ Số tiết KN 1:", min_value=0, value=0, step=1)
    with kn_col3:
        loai_kiem_nhiem_2 = st.text_input("📝 Kiêm nhiệm 2:", placeholder="VD: Công đoàn, Thư viện...")
    with kn_col4:
        kiem_nhiem_2 = st.number_input("⏱️ Số tiết KN 2:", min_value=0, value=0, step=1)

    if selected_teacher != "-- Chọn giáo viên --" and up_ppct:
        if ('report_data' not in st.session_state or
                st.session_state.get('current_teacher') != selected_teacher or
                st.session_state.get('current_week') != selected_week):
            st.session_state.report_data = []
            st.session_state.current_teacher = selected_teacher
            st.session_state.current_week = selected_week
            st.session_state.df_ppct = None

        if st.button("🚀 TẠO BÁO CÁO THEO TUẦN", use_container_width=True, type="primary"):
            # Xử lý tất cả các file PPCT được up lên và gộp chung lại
            all_dfs = []
            has_error = False

            for f in up_ppct:
                if f.name.endswith('.csv'):
                    df_raw = pd.read_csv(f)
                else:
                    df_raw = pd.read_excel(f)

                df_clean = clean_ppct_dataframe(df_raw)

                c_tiet = next((c for c in df_clean.columns if 'tiết' in str(c).lower()), None)
                c_bai = next((c for c in df_clean.columns if 'bài' in str(c).lower() or 'nội dung' in str(c).lower()),
                             None)
                c_lop = next((c for c in df_clean.columns if 'lớp' in str(c).lower() or 'khối' in str(c).lower()), None)
                c_mon = next((c for c in df_clean.columns if 'môn' in str(c).lower()), None)
                c_tuan = next((c for c in df_clean.columns if 'tuần' in str(c).lower()), None)

                if not c_tiet or not c_bai or not c_tuan:
                    st.error(f"❌ File '{f.name}' thiếu cột Tiết, Tên bài hoặc Tuần. Bỏ qua file này.")
                    has_error = True
                    continue

                # Chuẩn hóa dữ liệu để gộp an toàn
                temp_df = pd.DataFrame()
                temp_df['Tiết PPCT'] = pd.to_numeric(df_clean[c_tiet], errors='coerce')
                temp_df['Tên bài'] = df_clean[c_bai]
                temp_df['Tuần_Clean'] = pd.to_numeric(df_clean[c_tuan].astype(str).str.extract(r'(\d+)')[0],
                                                      errors='coerce')

                if c_lop:
                    temp_df['Khối_Clean'] = df_clean[c_lop].astype(str).str.extract(r'(\d+)')
                else:
                    temp_df['Khối_Clean'] = ""

                if c_mon:
                    temp_df['Môn_Clean'] = df_clean[c_mon].astype(str).str.lower().str.strip()
                else:
                    # Mẹo: Dùng tên file làm môn học nếu file không có cột Môn (VD: "Báo giảng GDĐP.csv" -> "gdđp")
                    inferred_mon = f.name.split('.')[0].lower().replace('báo giảng', '').strip()
                    temp_df['Môn_Clean'] = inferred_mon

                all_dfs.append(temp_df)

            if not all_dfs:
                st.error("❌ Không có file PPCT nào hợp lệ để xử lý!")
            else:
                df_ppct = pd.concat(all_dfs, ignore_index=True)
                st.session_state.df_ppct = df_ppct

                teacher_lessons = teachers_dict[selected_teacher]

                day_map = {'Thứ 2': 2, 'Thứ 3': 3, 'Thứ 4': 4, 'Thứ 5': 5, 'Thứ 6': 6, 'Thứ 7': 7, 'Chủ Nhật': 8}
                buoi_map = {'Sáng': 1, 'Chiều': 2}
                teacher_lessons.sort(key=lambda x: (day_map.get(x['Thứ'], 9), buoi_map.get(x['Buổi'], 3), x['Tiết']))

                final_report_data = []
                occurrence_tracker = {}

                for ls in teacher_lessons:
                    lop_tkb = str(ls['Lớp'])
                    mon = str(ls['Môn'])

                    match_khoi = re.search(r'\d+', lop_tkb)
                    khoi = match_khoi.group() if match_khoi else lop_tkb
                    mon_tkb_clean = mon.lower().strip()

                    key_occurrence = f"{lop_tkb}_{mon_tkb_clean}"
                    occurrence_tracker[key_occurrence] = occurrence_tracker.get(key_occurrence, 0) + 1
                    current_idx = occurrence_tracker[key_occurrence] - 1

                    query = (df_ppct['Tuần_Clean'] == selected_week)

                    # Sử dụng bộ lọc đã được tối ưu
                    mask_khoi = (df_ppct['Khối_Clean'] == khoi) | (df_ppct['Khối_Clean'] == "") | df_ppct[
                        'Khối_Clean'].isna()
                    query = query & mask_khoi

                    mask_mon = df_ppct['Môn_Clean'].apply(
                        lambda x: mon_tkb_clean in str(x) or str(x) in mon_tkb_clean or
                                  (mon_tkb_clean == 'tin' and str(x) == 'tin học') or
                                  (mon_tkb_clean == 'tin học' and str(x) == 'tin')
                    )
                    query = query & mask_mon

                    match_df = df_ppct[query].sort_values(by='Tiết PPCT')

                    if current_idx < len(match_df):
                        tiet_hien_tai = int(match_df.iloc[current_idx]['Tiết PPCT'])
                        ten_bai = str(match_df.iloc[current_idx]['Tên bài']).strip()
                    else:
                        tiet_hien_tai = 0
                        ten_bai = f"⚠️ Tuần {selected_week} PPCT chỉ có {len(match_df)} tiết cho khối này"

                    final_report_data.append({
                        "Thứ": ls['Thứ'], "Buổi": ls['Buổi'], "Tiết": ls['Tiết'],
                        "Lớp": lop_tkb, "Môn": mon, "Tiết PPCT": tiet_hien_tai, "Tên Bài": ten_bai,
                        "Khối": khoi,
                        "Loại Tiết": "Thực dạy / Kiêm nhiệm"  # Mặc định
                    })

                st.session_state.report_data = final_report_data

        if st.session_state.report_data:
            st.markdown(f"### 📋 Xem trước Báo Cáo (Tuần {selected_week})")

            # --- Thêm khu vực bổ sung tiết ngoài TKB ---
            with st.expander("➕ Bổ sung tiết ngoài TKB (Dạy bù, dạy thay, bồi dưỡng...)", expanded=False):
                st.write("Nhập thông tin tiết dạy bổ sung:")
                c1, c2, c3, c4 = st.columns(4)
                add_thu = c1.selectbox("Thứ", ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ Nhật"])
                add_buoi = c2.selectbox("Buổi", ["Sáng", "Chiều"])
                add_tiet = c3.number_input("Tiết", min_value=1, max_value=15, value=1)
                add_lop = c4.selectbox("Lớp", options=list_all_lops)

                c5, c6, c7, c8 = st.columns(4)
                add_mon = c5.selectbox("Môn", options=list_all_mons)
                add_tiet_ppct = c6.number_input("Tiết PPCT (Thêm mới)", min_value=0, value=0)
                add_loai_tiet = c7.selectbox("Loại Tiết (Thêm mới)", [
                    "Thực dạy / Kiêm nhiệm",
                    "Đi công tác",
                    "Dạy thay",
                    "Lấp giờ, tăng tiết, bù",
                    "Coi KT, dự giờ, BD, PĐ"
                ], index=3)  # Mặc định để mục Lấp giờ, tăng bù

                # Tự động nhảy Tên bài dựa trên Tiết PPCT
                auto_ten_bai = ""
                if add_tiet_ppct > 0 and st.session_state.df_ppct is not None:
                    match_khoi = re.search(r'\d+', add_lop)
                    khoi = match_khoi.group() if match_khoi else add_lop
                    mon_clean = str(add_mon).lower().strip()

                    found_name = find_lesson_name(st.session_state.df_ppct, khoi, add_tiet_ppct, mon_clean)
                    if "⚠️" not in found_name:
                        auto_ten_bai = found_name

                add_ten_bai = c8.text_input("Tên bài / Nội dung", value=auto_ten_bai, placeholder="VD: Ôn tập học kỳ")

                if st.button("Thêm tiết này vào báo cáo"):
                    match_khoi = re.search(r'\d+', add_lop)
                    khoi = match_khoi.group() if match_khoi else add_lop

                    st.session_state.report_data.append({
                        "Thứ": add_thu,
                        "Buổi": add_buoi,
                        "Tiết": add_tiet,
                        "Lớp": add_lop,
                        "Môn": add_mon,
                        "Tiết PPCT": int(add_tiet_ppct),
                        "Tên Bài": add_ten_bai if add_ten_bai else auto_ten_bai,
                        "Khối": khoi,
                        "Loại Tiết": add_loai_tiet
                    })

                    # Sắp xếp lại danh sách tự động theo thứ, buổi, tiết để dễ xem và hỗ trợ Gộp Ô (Merge cell)
                    day_map = {'Thứ 2': 2, 'Thứ 3': 3, 'Thứ 4': 4, 'Thứ 5': 5, 'Thứ 6': 6, 'Thứ 7': 7, 'Chủ Nhật': 8}
                    buoi_map = {'Sáng': 1, 'Chiều': 2}
                    st.session_state.report_data.sort(key=lambda x: (
                        day_map.get(x['Thứ'], 9),
                        buoi_map.get(x['Buổi'], 3),
                        int(x['Tiết']) if str(x['Tiết']).isdigit() else 99
                    ))

                    st.rerun()

            st.info(
                "💡 Bảng dưới đây đã được nâng cấp thành các **Menu xổ xuống**. Việc chọn Tiết PPCT và Loại tiết sẽ diễn ra cực kỳ nhanh chóng. Tên bài sẽ tự động nhảy theo ngay lập tức. Nhấn nút ❌ để xóa tiết dạy nghỉ lễ.")

            st.markdown("---")
            header_cols = st.columns([2.5, 1.5, 4, 2.5, 0.5])
            header_cols[0].markdown("**Lớp - Môn (Thời gian)**")
            header_cols[1].markdown("**Tiết PPCT**")
            header_cols[2].markdown("**Tên Bài / Nội dung**")
            header_cols[3].markdown("**Loại Tiết**")
            header_cols[4].markdown("**Xóa**")

            idx_to_remove = None
            for i, row in enumerate(st.session_state.report_data):
                cols = st.columns([2.5, 1.5, 4, 2.5, 0.5])

                # Cột 1: Thông tin tĩnh
                thu_buoi = f"{row['Thứ']}, {row['Buổi']} (T{row['Tiết']})"
                cols[0].markdown(
                    f"<div style='padding-top: 5px; font-size: 14px;'><b>{row['Lớp']} - {row['Môn']}</b><br/><span style='color: gray;'>{thu_buoi}</span></div>",
                    unsafe_allow_html=True)

                # Cột 2: Tiết PPCT (Menu xổ xuống)
                tiet_options = list(range(0, 151))  # Hỗ trợ tối đa 150 tiết/năm
                try:
                    current_tiet = int(row['Tiết PPCT'])
                except:
                    current_tiet = 0

                if current_tiet not in tiet_options:
                    tiet_options.append(current_tiet)
                    tiet_options.sort()

                new_tiet = cols[1].selectbox("Tiết PPCT", options=tiet_options, index=tiet_options.index(current_tiet),
                                             key=f"tiet_{i}", label_visibility="collapsed")

                # Cập nhật Tên Bài ngay nếu Tiết PPCT thay đổi
                display_bai = str(row['Tên Bài'])
                if new_tiet != current_tiet:
                    st.session_state.report_data[i]['Tiết PPCT'] = new_tiet
                    khoi = row['Khối']
                    mon_clean = str(row['Môn']).lower().strip()

                    display_bai = find_lesson_name(st.session_state.df_ppct, khoi, new_tiet, mon_clean)
                    st.session_state.report_data[i]['Tên Bài'] = display_bai

                # Cột 3: Tên bài (Cho phép gõ chữ nếu muốn tự sửa tay)
                new_bai = cols[2].text_input("Tên Bài", value=display_bai, key=f"bai_{i}", label_visibility="collapsed")
                if new_bai != display_bai:
                    st.session_state.report_data[i]['Tên Bài'] = new_bai

                # Cột 4: Loại Tiết
                loai_options = [
                    "Thực dạy / Kiêm nhiệm",
                    "Đi công tác",
                    "Dạy thay",
                    "Lấp giờ, tăng tiết, bù",
                    "Coi KT, dự giờ, BD, PĐ"
                ]
                current_loai = row.get('Loại Tiết', "Thực dạy / Kiêm nhiệm")
                loai_idx = loai_options.index(current_loai) if current_loai in loai_options else 0
                new_loai = cols[3].selectbox("Loại Tiết", options=loai_options, index=loai_idx, key=f"loai_{i}",
                                             label_visibility="collapsed")

                if new_loai != current_loai:
                    st.session_state.report_data[i]['Loại Tiết'] = new_loai

                # Cột 5: Nút xóa
                if cols[4].button("❌", key=f"del_{i}", help="Xóa tiết này"):
                    idx_to_remove = i

            # Thực hiện xóa nếu có tiết được chọn
            if idx_to_remove is not None:
                st.session_state.report_data.pop(idx_to_remove)
                st.rerun()

            missing_ppct = any("⚠️" in str(row["Tên Bài"]) for row in st.session_state.report_data)
            if missing_ppct:
                st.warning(
                    "⚠️ Vẫn còn một số tiết bị báo lỗi ⚠️. Có thể PPCT tuần này thiếu bài, hãy sửa lại Tiết PPCT ở bảng trên.")

            output_file = create_excel_report(selected_teacher, chuc_vu, to_chuyen_mon, nam_hoc, hoc_ky, selected_week,
                                              start_date, end_date, loai_kiem_nhiem, kiem_nhiem, loai_kiem_nhiem_2,
                                              kiem_nhiem_2, st.session_state.report_data)

            with open(output_file, "rb") as f:
                st.download_button(
                    label=f"📥 TẢI FILE BÁO CÁO TUẦN {selected_week}",
                    data=f, file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )

            st.markdown(
                "<div style='text-align: center; margin-top: 15px; color: gray;'>@copyright Đỗ Đặng Toàn@</div>",
                unsafe_allow_html=True)